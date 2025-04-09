import pdfplumber
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# File paths
#pdf_path=pdf_path ="D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/X2R1/X2R1-T5.3-D-SIE-102-20_-_D5.1_-_Moving_Block_System_Requirements.pdf"
#pdf_path ="D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/X2R3/X2R3-T4_3-D-SMD-008-19_-_D4.2Part3-SystemSpecification.pdf"
#pdf_path = "D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/X2R5/X2R5-T4_2-D-SMD-003-23_-_D41Part3SystemSpecification.pdf"
pdf_path ="D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/X2R1/X2R1-T5.3-D-SIE-102-20_-_D5.1_-_Moving_Block_System_Requirements.pdf"
excel_template = "D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/req Eng.xlsx"

# Output file
pdf_filename = os.path.basename(pdf_path).replace(".pdf", "")
output_excel = f"C:/Users/aroua/Downloads/{pdf_filename}_Result.xlsx"

# Load Excel template
wb = load_workbook(excel_template)
ws = wb.active

# Extract column headers from template
columns_needed = [cell.value for cell in ws[1] if cell.value]
required_columns = ["Topic", "Requirement ID", "Description", "Traceability"]

# Patterns for extraction
#topic_pattern = re.compile(r"^\s*(\d+\.\d+)\s+([A-Za-z][A-Za-z0-9 \-]+)")
req_pattern = re.compile(r"(REQ-[A-Za-z0-9]+-\d+|\bREQ-[A-Za-z0-9]+)\s*(\[[^\]]+\])?")
traceability_pattern = re.compile(r"\[(X2R\d+ D\d+\.\d+: REQ-[A-Za-z0-9-]+)\]")

footer_pattern = re.compile(r"(GA\s*\d+\s*)?Page\s+\d+\s+of\s+\d+", re.IGNORECASE)


# Function to extract description
def extract_description(lines, start_idx):
    """Extract the requirement description."""
    description = []
    for i in range(start_idx, len(lines)):
        if "Rationale:" in lines[i] or "Guidance:" in lines[i]:  # Stop at Rationale/Guidance
            break
        clean_line = footer_pattern.sub("", lines[i]).strip()  # Remove footer
        if clean_line:
            description.append(clean_line)
    return "\n".join(description).strip()  # Keep structured formatting

# Function to extract requirements
def extract_requirements(pdf_path):
    """Extract requirements from the PDF."""
    requirements = []
    current_topic = "Unknown"
    last_traceability = "[Not Provided]"  # Default traceability if none found

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True)
            if text:
                lines = text.split("\n")

                for idx, line in enumerate(lines):
                    # Remove footers
                    line = footer_pattern.sub("", line).strip()

                    # Try normal topic format: "6.1 Train Location"
                    normal_topic = re.match(r"^\s*(\d+\.\d+)\s+([A-Za-z][A-Za-z0-9 \-/]+)$", line)
                    compact_topic = re.match(r"^\s*(\d+\.\d+)([A-Z][A-Za-z0-9]+)\s*(.*)$", line)

                    if normal_topic:
                        section_number = normal_topic.group(1)
                        topic_name = normal_topic.group(2).strip()
                        current_topic = f"{section_number} - {topic_name}"
                        print(f"✅ Detected Topic (normal): {current_topic}")

                    elif compact_topic:
                        section_number = compact_topic.group(1)
                        section_code = compact_topic.group(2)
                        rest = compact_topic.group(3).strip()
                        current_topic = f"{section_number} - {section_code} {rest}".strip()
                        print(f"✅ Detected Topic (compact): {current_topic}")

                    # Detect traceability anywhere in the text
                    traceability_match = traceability_pattern.search(line)
                    if traceability_match:
                        last_traceability = traceability_match.group(1).strip()
                        print(f"🔵 Found Traceability: {last_traceability}")
                    elif "[New]" in line:  # Special case for "[New]" traceability
                        last_traceability = "[New]"
                        print("🔵 Found Traceability: New")

                    # Extract requirement ID
                    req_match = req_pattern.search(line)
                    if req_match:
                        req_id = req_match.group(1)

                        # Détection du bon format de traceability
                        if req_match.group(2):  # If traceability is on the same line
                            traceability = req_match.group(2).strip("[]")
                        else:
                            traceability = last_traceability  # Otherwise, use the last traceability found

                        # Extract description
                        description = extract_description(lines, idx + 1)

                        print(f"📌 Storing: {req_id} | {current_topic} | {traceability}")

                        # Append to requirements list
                        requirements.append((current_topic, req_id, description, traceability))

    return requirements

# Extract requirements
data = extract_requirements(pdf_path)

# Convert to DataFrame
extracted_df = pd.DataFrame(data, columns=required_columns)
# Remove duplicate Requirement IDs (keep the first occurrence)
extracted_df.drop_duplicates(subset=["Requirement ID"], keep="first", inplace=True)


# ✅ Vérification des premières lignes
print(extracted_df.head())

# Assurez-vous que toutes les colonnes nécessaires existent
for col in columns_needed:
    if col not in extracted_df.columns:
        extracted_df[col] = ""  # Remplissage des colonnes manquantes

# Écriture des données dans l'Excel
for row_idx, row in extracted_df.iterrows():
    for col_idx, col_name in enumerate(required_columns):
        ws[f"{chr(65 + col_idx)}{row_idx + 2}"] = row[col_name]

# Mise en forme : Retour automatique à la ligne et ajustement des colonnes
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(required_columns)):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)  # Wrap text

# Ajustement automatique de la hauteur des lignes
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):  # Seulement pour Description
    for cell in row:
        ws.row_dimensions[cell.row].height = None

# Ajustement de la largeur des colonnes (limité à 50 caractères)
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Get column letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = min(max_length + 5, 50)

# Enregistrer le fichier final
wb.save(output_excel)

print(f"✅ Extraction completed! File saved at {output_excel}")
